import sys
import random
import openpyxl
import json
import math
import os
from PyQt6.QtWidgets import QApplication, QMainWindow
from PyQt6.QtGui import QPixmap, QPainter
from PyQt6.QtCore import Qt, QTimer


def distance(point1, point2):
    """
    Calcule la distance entre deux points.

    Args:
        point1 (tuple): Coordonnées du premier point.
        point2 (tuple): Coordonnées du deuxième point.

    Returns:
        float: La distance entre les deux points.
    """
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)


class Carte(QMainWindow):
    MARGIN = 10  # Marge autour de la carte
    SCREEN_WIDTH = 1033  # Largeur de la fenêtre de la carte
    SCREEN_HEIGHT = 580  # Hauteur de la fenêtre de la carte
    GRID_SIZE = 20  # Taille d'une grille de la carte

    def __init__(self):
        super().__init__()
        self.setWindowTitle("POKEMON")
        # Taille minimale de la fenêtre
        self.setMinimumSize(Carte.SCREEN_WIDTH, Carte.SCREEN_HEIGHT)

        # Chemin du dossier contenant les images d'arrière-plan
        background_folder = "../Images/background/"

        # Liste pour stocker les chemins d'accès aux images d'arrière-plan
        self.background_images = []

        # Parcours des fichiers dans le dossier
        for filename in os.listdir(background_folder):
            # Vérifie si le fichier a l'extension .jpeg
            if filename.endswith(".jpeg"):
                # Ajoute le chemin d'accès complet à la liste
                self.background_images.append(os.path.join(background_folder, filename))

        # Chargement initial de l'image d'arrière-plan
        self.background_index = 0
        self.background_image = QPixmap(self.background_images[self.background_index])

        # Création du joueur
        coord_surete = [(13, 5.25), (13, 9), (4, 7.25),
                        (6, 2), (25, 0.75), (35, 0.75)]
        coord_choice = coord_surete[random.randint(0, len(coord_surete) - 1)]
        self.joueur = Joueur(coord_choice[0] / 2, 2 * coord_choice[1])

        # Création des Pokemons
        self.lespokemons = LesPokemons()
        self.pokemons_libres = self.lespokemons.extract_pokemons_libres(10)

        self.pokemon_images = {}
        for pokemon_name in self.lespokemons.pokemons:
            image_path = f"../Images/Pokemon_images/{pokemon_name}.PNG"
            self.pokemon_images[pokemon_name] = QPixmap(image_path)

        # Minuterie pour changer l'image de l'arrière-plan toutes les 5 secondes
        self.background_timer = QTimer(self)
        self.background_timer.timeout.connect(self.change_background)
        self.background_timer.start(5000)  # Change l'image toutes les 5000 ms (5 secondes)

    def change_background(self):
        """
        Change l'image de l'arrière-plan toutes les 5 secondes.
        """
        self.background_index = (self.background_index + 1) % len(self.background_images)
        self.background_image = QPixmap(self.background_images[self.background_index])
        self.update()  # Redessine la fenêtre pour afficher la nouvelle image

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.drawPixmap(self.rect(), self.background_image)

        x = round(self.joueur.row * (self.width() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        y = round(self.joueur.col * (self.height() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        painter.drawPixmap(x, y, self.joueur.sizeJ, self.joueur.sizeJ, self.joueur.image)

        echelleX = (self.width() - 2 * Carte.MARGIN) / 40
        echelleY = (self.height() - 2 * Carte.MARGIN) / 10

        for pokemon_name, coord in self.pokemons_libres.items():
            X = round(coord['x'] * echelleX)
            Y = round(coord['y'] * echelleY)
            painter.drawPixmap(X, Y, self.lespokemons.sizeP, self.lespokemons.sizeP, self.pokemon_images[pokemon_name])

        for pokemon_name, coord in self.lespokemons.pokemons.items():
            if not self.pokemons_libres.get(pokemon_name) and coord['visible']:
                X = round(coord['x'] * echelleX)
                Y = round(coord['y'] * echelleY)
                painter.drawPixmap(X, Y, self.lespokemons.sizeP, self.lespokemons.sizeP,
                                   self.pokemon_images[pokemon_name])

    def distance_joueur_pokemons(self):
        distances = {}
        joueur_position = (2 * self.joueur.row, self.joueur.col / 2)
        for pokemon_name, info in self.lespokemons.pokemons.items():
            pokemon_position = (info['x'], info['y'])
            dist = distance(joueur_position, pokemon_position)
            distances[pokemon_name] = dist
        return distances

    def keyPressEvent(self, event):
        """
            Gère les événements de pression de touche.

            Args:
                event: L'événement de pression de touche.

            """
        # Déplace le joueur en fonction de la touche pressée

        if event.key() == Qt.Key.Key_Up:
            self.joueur.deplacer("haut")
        elif event.key() == Qt.Key.Key_Down:
            self.joueur.deplacer("bas")
        elif event.key() == Qt.Key.Key_Left:
            self.joueur.deplacer("gauche")
        elif event.key() == Qt.Key.Key_Right:
            self.joueur.deplacer("droite")
        # Affiche le Pokémon le plus proche du joueur
        self.lespokemons.show_nearest((2 * self.joueur.row, self.joueur.col / 2))
        # Actualise l'affichage de la carte
        self.repaint()


class Joueur:
    def __init__(self, row, col):
        self.row = row
        self.col = col
        self.sizeJ = 40
        self.image = QPixmap("../Images/Pion.png")

    def deplacer(self, direction, step=0.5):
        if direction == "haut" and self.col > 0:
            self.col -= step
        elif direction == "bas" and self.col < Carte.GRID_SIZE - 1:
            self.col += step
        elif direction == "gauche" and self.row > 0:
            self.row -= step
        elif direction == "droite" and self.row < Carte.GRID_SIZE - 1:
            self.row += step


class LesPokemons:
    def __init__(self):
        super().__init__()
        self.sizeP = 20
        self.pokemons = {}
        wb = openpyxl.load_workbook("../Donnees_crees/pokemon_coordinates_P.xlsx")
        sheet = wb.active
        selected_rows = random.sample(range(1, sheet.max_row + 1), 30)
        for row in selected_rows:
            cell1 = sheet.cell(row, 1)
            cell2 = sheet.cell(row, 2)
            pokemon_name = cell1.value
            x, y = json.loads(cell2.value)
            self.pokemons[pokemon_name] = {
                'x': x,
                'y': y,
                'visible': False
            }

    def extract_pokemons_libres(self, num_elements):
        if num_elements > len(self.pokemons):
            num_elements = min(num_elements, len(self.pokemons))
        selected_keys = random.sample(list(self.pokemons.keys()), num_elements)
        extracted_elements = {key: self.pokemons[key] for key in selected_keys}
        return extracted_elements

    def show_nearest(self, joueur_position):
        nearest_pokemon = None
        nearest_distance = 1
        for pokemon_name, info in self.pokemons.items():
            dist = distance(joueur_position, (info['x'], info['y']))
            if dist < nearest_distance:
                nearest_distance = dist
                nearest_pokemon = pokemon_name
        if nearest_pokemon:
            for pokemon_name, info in self.pokemons.items():
                if pokemon_name == nearest_pokemon:
                    info['visible'] = True
                else:
                    info['visible'] = False


# Programme Principal
def main():
    app = QApplication(sys.argv)
    window = Carte()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
