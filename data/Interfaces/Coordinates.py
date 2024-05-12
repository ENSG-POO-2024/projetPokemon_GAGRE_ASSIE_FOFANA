import sys
import random
import openpyxl
import json
import math
from PyQt6.QtWidgets import QApplication, QMainWindow
from PyQt6.QtGui import QPixmap, QPainter, QColor
from PyQt6.QtCore import Qt


def distance(point1, point2):
    """
    Calcule la distance entre deux points.
    """
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)


class Carte(QMainWindow):
    MARGIN = 10  # Marge autour de la carte
    SCREEN_WIDTH = 800  # Largeur de la fenêtre de la carte
    SCREEN_HEIGHT = 800  # Hauteur de la fenêtre de la carte
    CordSupX = 100  # Coordonnée X supérieure
    CordSupY = 100  # Coordonnée Y supérieure
    GRID_SIZE = 20  # Taille de la grille de la carte

    def __init__(self):
        super().__init__()
        self.setWindowTitle("POKEMON")
        # Taille minimale de la fenêtre
        self.setMinimumSize(Carte.SCREEN_WIDTH, Carte.SCREEN_HEIGHT)

        # Charger l'image d'arrière-plan
        self.background_image = QPixmap("../Images/background/back.jpg")

        # Création du joueur

        # Définition des coordonnées sûres où le joueur peut apparaître initialement
        coord_surete = [(13, 5.25), (13, 9), (4, 7.25),
                        (6, 2), (25, 0.75), (35, 0.75)]
        # Choix aléatoire d'une paire de coordonnées parmi celles définies dans coord_surete
        coord_choice = coord_surete[random.randint(0, len(coord_surete) - 1)]
        # Les coordonnées sont ajustées avant d'être passées pour positionner le joueur correctement sur la carte
        self.joueur = Joueur(coord_choice[0] / 2, 2 * coord_choice[1])

        # Création des Pokemons

        self.lespokemons = LesPokemons()
        # Extraction des pokemons libres
        # L'argument 10 spécifie le nombre de pokemons libres à extraire
        self.pokemons_libres = self.lespokemons.extract_pokemons_libres(10)

        # Chargement des images des Pokémons
        self.pokemon_images = {}
        for pokemon_name in self.lespokemons.pokemons:
            image_path = f"../Images/Pokemon_images/{pokemon_name}.PNG"
            self.pokemon_images[pokemon_name] = QPixmap(image_path)

    def paintEvent(self, event):
        """
        Elle la carte, le joueur et les pokemons.
        """
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.drawPixmap(self.rect(), self.background_image)

        # Dessin du joueur
        x = round(self.joueur.row * (self.width() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        y = round(self.joueur.col * (self.height() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        painter.drawPixmap(x, y, self.joueur.cell_size, self.joueur.cell_size, self.joueur.image)

        # Dessin des pokemons libres
        for pokemon_name, info in self.pokemons_libres.items():
            X = round(info['x'] * 19.5 + Carte.MARGIN)
            Y = round(info['y'] * 78 + Carte.MARGIN)
            painter.drawPixmap(X, Y, self.joueur.cell_size, self.joueur.cell_size, self.pokemon_images[pokemon_name])

        # Dessin des pokemons sauvages
        for pokemon_name, info in self.lespokemons.pokemons.items():
            if not self.pokemons_libres.get(pokemon_name) and info['visible']:
                X = round(info['x'] * 19.5 + Carte.MARGIN)
                Y = round(info['y'] * 78 + Carte.MARGIN)
                painter.drawPixmap(X, Y, self.joueur.cell_size, self.joueur.cell_size,
                                   self.pokemon_images[pokemon_name])

    def distance_joueur_pokemons(self):
        """
        Calcule la distance entre le joueur et tous les Pokémon.
        Retourne un dictionnaire avec les distances pour chaque Pokémon.
        """
        distances = {}
        joueur_position = (2 * self.joueur.row, self.joueur.col / 2)
        for pokemon_name, info in self.lespokemons.pokemons.items():
            pokemon_position = (info['x'], info['y'])
            dist = distance(joueur_position, pokemon_position)
            print(dist)
            distances[pokemon_name] = dist
        return distances

    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Up:
            self.joueur.deplacer("haut")
        elif event.key() == Qt.Key.Key_Down:
            self.joueur.deplacer("bas")
        elif event.key() == Qt.Key.Key_Left:
            self.joueur.deplacer("gauche")
        elif event.key() == Qt.Key.Key_Right:
            self.joueur.deplacer("droite")
        self.lespokemons.show_nearest((2 * self.joueur.row, self.joueur.col / 2))
        self.repaint()


class Joueur:
    COLOR = QColor(255, 0, 0)  # Couleur rouge pour le joueur

    def __init__(self, row, col):
        self.row = row
        self.col = col
        self.cell_size = 40
        self.image = QPixmap("../Images/Pion.png")  # Chargement de l'image du joueur

    def deplacer(self, direction, step=0.5):
        """
        Déplace le joueur sur la carte en vérifiant les limites.
        """
        if direction == "haut" and self.col > 0:
            self.col -= step
        elif direction == "bas" and self.col < Carte.GRID_SIZE - 1:
            self.col += step
        elif direction == "gauche" and self.row > 0:
            self.row -= step
        elif direction == "droite" and self.row < Carte.GRID_SIZE - 1:
            self.row += step


class LesPokemons:
    COLORS = QColor(0, 0, 0)  # Couleur noire pour les Pokémons sauvages
    COLORL = QColor(255, 255, 0)  # Couleur blanche pour les Pokémons libres

    def __init__(self):
        super().__init__()
        self.pokemons = {}
        wb = openpyxl.load_workbook("../Brutes/pokemon_coordinates.xlsx")
        sheet = wb.active
        # Sélectionner un échantillon aléatoire de 30 éléments parmi tous les éléments du fichier Excel
        selected_rows = random.sample(range(1, sheet.max_row + 1), 30)
        for row in selected_rows:
            cell1 = sheet.cell(row, 1)
            cell2 = sheet.cell(row, 2)
            pokemon_name = cell1.value
            x, y = json.loads(cell2.value)
            self.pokemons[pokemon_name] = {
                'x': x,
                'y': y,
                'visible': False
            }

    def extract_pokemons_libres(self, num_elements):
        if num_elements > len(self.pokemons):
            num_elements = min(num_elements, len(self.pokemons))
        selected_keys = random.sample(list(self.pokemons.keys()), num_elements)
        extracted_elements = {key: self.pokemons[key] for key in selected_keys}
        return extracted_elements

    def show_nearest(self, joueur_position):
        nearest_pokemon = None
        nearest_distance = 1  # float("Inf")
        for pokemon_name, info in self.pokemons.items():
            dist = distance(joueur_position, (info['x'], info['y']))
            if dist < nearest_distance:
                nearest_distance = dist
                nearest_pokemon = pokemon_name

        if nearest_pokemon:
            for pokemon_name, info in self.pokemons.items():
                if pokemon_name == nearest_pokemon:
                    info['visible'] = True
                else:
                    info['visible'] = False


def main():
    app = QApplication(sys.argv)
    window = Carte()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
