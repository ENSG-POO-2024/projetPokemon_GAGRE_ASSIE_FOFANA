import sys
import random
import openpyxl
import json
import math
from PyQt6.QtWidgets import QApplication, QMainWindow, QLabel
from PyQt6.QtGui import QPixmap
from PyQt6.QtGui import QPainter, QColor, QFont
from PyQt5.QtCore import Qt


# from PyQt6.QtCore import Qt


def distance(point1, point2):
    """
    Calcule la distance entre deux points.
    """
    return math.sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)


class Carte(QMainWindow):
    MARGIN = 10
    SCREEN_WIDTH = 800
    SCREEN_HEIGHT = 800
    CordSupX = 100
    CordSupY = 100
    GRID_SIZE = 20

    def __init__(self):
        super().__init__()
        self.setWindowTitle("POKEMON")
        self.setMinimumSize(Carte.SCREEN_WIDTH, Carte.SCREEN_HEIGHT)

        # Charger l'image d'arrière-plan
        self.background_image = QPixmap("Images/BackgroundImage.jpeg")

        # Création du joueur
        rows = random.randint(0, Carte.GRID_SIZE)
        cols = random.randint(0, Carte.GRID_SIZE)
        self.joueur = Joueur(rows, cols)

        # Création des Pokemons
        self.lespokemons = LesPokemons()

        # Ajout des labels pour afficher les noms des Pokémons
        self.labels = {}
        for pokemon_name, info in self.lespokemons.pokemons.items():
            label = QLabel(pokemon_name, self)
            label.setStyleSheet("color: black")  # Changer la couleur du texte
            #label.setAlignment(Qt.AlignCenter)  # Aligner le texte au centre
            # Créer une nouvelle police avec une taille plus grande
            font = QFont()
            font.setPointSize(16)  # Changer 16 par la taille de police souhaitée
            # Appliquer la nouvelle police au QLabel
            label.setFont(font)
            label.setGeometry(round(info[0] * 19.5), round(info[1] * 78), 100, 20)  # Position et taille du label
            self.labels[pokemon_name] = label

    def paintEvent(self, event):
        """
        Elle la carte, le joueur et les pokemons.
        """
        # Dessin de la carte

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        painter.drawPixmap(self.rect(), self.background_image)

        # Dessin du joueur

        painter.setPen(Joueur.COLOR)
        painter.setBrush(Joueur.COLOR)
        # Calculer les coordonnées du coin supérieur gauche de l'ellipse
        x = round(self.joueur.row * (self.width() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        y = round(self.joueur.col * (self.height() - 2 * Carte.MARGIN) // Carte.GRID_SIZE + Carte.MARGIN)
        # Dessiner l'ellipse centrée sur les coordonnées calculées
        painter.drawEllipse(x, y, self.joueur.cell_size, self.joueur.cell_size)

        # Dessiner les pokemons
        rayon = 10
        for pokemon in self.lespokemons.positionner():
            coord = pokemon[1]
            coordX = round(coord[0])
            coordY = round(coord[1])
            painter.setPen(LesPokemons.COLOR)
            painter.setBrush(LesPokemons.COLOR)
            painter.drawEllipse(coordX, coordY, rayon, rayon)

    def distance_joueur_pokemons(self):
        """
        Calcule la distance entre le joueur et tous les Pokémon.
        Retourne un dictionnaire avec les distances pour chaque Pokémon.
        """
        distances = {}
        joueur_position = (2*self.joueur.row, self.joueur.col/2)
        print('*********', joueur_position, '*********')
        for pokemon_name, (pokemon_x, pokemon_y) in self.lespokemons.pokemons.items():
            pokemon_position = (pokemon_x, pokemon_y)
            dist = distance(joueur_position, pokemon_position)
            distances[pokemon_name] = dist
        return distances

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Up:
            self.joueur.deplacer("haut")
        elif event.key() == Qt.Key_Down:
            self.joueur.deplacer("bas")
        elif event.key() == Qt.Key_Left:
            self.joueur.deplacer("gauche")
        elif event.key() == Qt.Key_Right:
            self.joueur.deplacer("droite")
        self.repaint()  # Redessine la fenêtre après le déplacement du joueur

        # Mettre à jour les distances entre le joueur et les pokemons
        distances = self.distance_joueur_pokemons()
        print(distances)


class Joueur:
    COLOR = QColor(255, 0, 0)  # On choisit une couleur pour le joueur : Rouge

    def __init__(self, row, col):
        self.row = row
        self.col = col
        self.cell_size = 20

    def deplacer(self, direction, step=0.0625):
        """
        Elle permet de déplacer le joueur sur la grille ou la carte en vérifiant que le joueur ne sort pas de la zone
        du jeu. direction: la direction dans laquelle le joueur se déplace
        """
        # Logique de déplacement
        if direction == "haut" and self.col > 0:
            self.col -= step
        elif direction == "bas" and self.col < Carte.GRID_SIZE - 1:
            self.col += step
        elif direction == "gauche" and self.row > 0:
            self.row -= step
        elif direction == "droite" and self.row < Carte.GRID_SIZE - 1:
            self.row += step
        elif direction == "haut-gauche" and self.col > 0 and self.row > 0:
            self.col -= step
            self.row -= step
        elif direction == "haut-droite" and self.col > 0 and self.row < Carte.GRID_SIZE - 1:
            self.col -= step
            self.row += step
        elif direction == "bas-gauche" and self.col < Carte.GRID_SIZE - 1 and self.row > 0:
            self.col += step
            self.row -= step
        elif direction == "bas-droite" and self.col < Carte.GRID_SIZE - 1 and self.row < Carte.GRID_SIZE - 1:
            self.col += step
            self.row += step


class LesPokemons:
    COLOR = QColor(0, 0, 0)  # Couleur noir pour les Pokémon

    def __init__(self):
        self.pokemons = {}  # Dictionnaire pour stocker les coordonnées des Pokémon
        # Chargement des coordonnées des Pokémon depuis le fichier Excel
        wb = openpyxl.load_workbook("essai.xlsx")
        sheet = wb.active
        for row in range(1, sheet.max_row + 1):
            cell1 = sheet.cell(row, 1)
            cell2 = sheet.cell(row, 2)
            pokemon_name = cell1.value
            x, y = json.loads(f"{cell2.value}")  # Convertir la valeur de la cellule en tuple
            self.pokemons[pokemon_name] = (x, y)

    def positionner(self):
        pokemons = []
        for pokemon_name, (x, y) in self.pokemons.items():
            X = x * 19.5 + Carte.MARGIN
            Y = y * 78 + Carte.MARGIN
            pokemons.append((pokemon_name, (X, Y)))
        return pokemons


def main():
    app = QApplication(sys.argv)
    window = Carte()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()