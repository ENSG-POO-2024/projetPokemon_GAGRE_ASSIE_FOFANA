import openpyxl
import json
wb = openpyxl.load_workbook("../pokemon_coordinates.xlsx")
sheet = wb.active
liste =[]
for row in range(1, sheet.max_row+1):
    cell1 = sheet.cell(row, 1)
    cell2 = sheet.cell(row, 2)
