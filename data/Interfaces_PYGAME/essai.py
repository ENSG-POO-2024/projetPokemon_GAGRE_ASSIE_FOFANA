import pygame
import random
import os
import openpyxl
import json


class Carte:
    # Initialisation de Pygame
    pygame.init()

    # Configuration de la fenêtre du jeu
    MARGIN = 10  # Taille de la marge autour de la carte
    SCREEN_WIDTH = 800
    SCREEN_HEIGHT = 800
    GRID_SIZE = 10  # Nombre de cellule de la grille sur une ligne ou sur une colonne.
    CELL_SIZE = (SCREEN_WIDTH - 2 * MARGIN) // GRID_SIZE  # Taille d'une cellule de la grille
    screen = pygame.display.set_mode((SCREEN_WIDTH, SCREEN_HEIGHT))  # Affiche l'écran principal
    pygame.display.set_caption("POKEMON")  # Affiche le titre donné en haut du dessin

    # Couleurs
    # J'ai chosit la couleur blanche
    WHITE = (255, 255, 255)
    BLACK = (0, 0, 0)

    # Chargement de l'image d'arrière-plan
    background_image = pygame.image.load("Images/Jardin.jpg")
    # Permet d'afficher l'image sur toute la surface de l'ecran en se limitant à la limite des marges.
    background_image = pygame.transform.scale(background_image, (SCREEN_WIDTH - MARGIN, SCREEN_HEIGHT - MARGIN))

    @classmethod
    def dessiner_grille(cls):
        """
        Elle dessine la carte ou grille.
        cls : cet paramètre fait référence à la classe elle même.
        Elle ne s'applique pas un objet de la classe Carte mais à la classe elle même.
        """
        # Effacer l'écran ou remplir l'ecran dans le but d'effacer ou cacher ce qui existait avant sur l'écran
        cls.screen.fill(cls.WHITE)

        # Dessiner l'image d'arrière-plan sur l'écran(screen) en spécifiant la coordonnée supérieur gauche
        cls.screen.blit(cls.background_image, (cls.MARGIN, cls.MARGIN),
                        pygame.Rect(cls.MARGIN, cls.MARGIN, cls.SCREEN_WIDTH - cls.MARGIN,
                                    cls.SCREEN_HEIGHT - cls.MARGIN))

        # Dessin du rectangle
        pygame.draw.rect(cls.screen, cls.BLACK,
                         pygame.Rect(cls.MARGIN, cls.MARGIN, cls.SCREEN_WIDTH - 2 * cls.MARGIN,
                                     cls.SCREEN_HEIGHT - 2 * cls.MARGIN), 1)

        # Dessiner le cadrillage
        for i in range(1, cls.GRID_SIZE):
            # Trace une droite horizonrale(ligne) à chaque boucle
            pygame.draw.line(cls.screen, cls.BLACK, (cls.MARGIN, cls.MARGIN + i * cls.CELL_SIZE),
                             (cls.SCREEN_WIDTH - cls.MARGIN, cls.MARGIN + i * cls.CELL_SIZE), 1)
            # Trace une droite verticale(colonne) à chaque boucle
            pygame.draw.line(cls.screen, cls.BLACK, (cls.MARGIN + i * cls.CELL_SIZE, cls.MARGIN),
                             (cls.MARGIN + i * cls.CELL_SIZE, cls.SCREEN_HEIGHT - cls.MARGIN), 1)

        # Rafraîchir l'écran
        pygame.display.flip()


class Joueur:
    def __init__(self, row, col):
        self.row = row
        self.col = col
        self.color = (255, 0, 0)  # On choisit une couleur pour le joueur : Rouge

    def deplacer(self, direction):
        """
        Elle permet de deplacer le joueur sur la grille ou la carte en vérifiant que le joueur ne sort pas de la zone
        du jeu. direction: la direction dans laquelle le joueur se deplace
        """
        # Calcul des limites des bords
        limit_left = Carte.MARGIN  # la limite gauche de la carte
        limit_right = Carte.SCREEN_WIDTH - Carte.MARGIN - Carte.CELL_SIZE  # la limite droite de la carte
        limit_top = Carte.MARGIN  # la limite haut de la carte
        limit_bottom = Carte.SCREEN_HEIGHT - Carte.MARGIN - Carte.CELL_SIZE  # la limite bas de la carte

        if direction == "haut" and self.col > 0:
            self.col -= 1
        elif direction == "bas" and self.col < Carte.GRID_SIZE - 1:
            self.col += 1
        elif direction == "gauche" and self.row > 0:
            self.row -= 1
        elif direction == "droite" and self.row < Carte.GRID_SIZE - 1:
            self.row += 1
        elif direction == "haut-gauche" and self.col > 0 and self.row > 0:
            self.col -= 1
            self.row -= 1
        elif direction == "haut-droite" and self.col > 0 and self.row < Carte.GRID_SIZE - 1:
            self.col -= 1
            self.row += 1
        elif direction == "bas-gauche" and self.col < Carte.GRID_SIZE - 1 and self.row > 0:
            self.col += 1
            self.row -= 1
        elif direction == "bas-droite" and self.col < Carte.GRID_SIZE - 1 and self.row < Carte.GRID_SIZE - 1:
            self.col += 1
            self.row += 1

        # Vérification des limites des bords
        self.row = max(limit_left // Carte.CELL_SIZE, min(self.row, limit_right // Carte.CELL_SIZE))
        self.col = max(limit_top // Carte.CELL_SIZE, min(self.col, limit_bottom // Carte.CELL_SIZE))

    def dessiner(self):
        """
        Elle dessine le joueur.
        """
        # Dessiner le joueur comme un cercle

        centre = (Carte.MARGIN + self.row * Carte.CELL_SIZE + Carte.CELL_SIZE // 2,
                  Carte.MARGIN + self.col * Carte.CELL_SIZE + Carte.CELL_SIZE // 2)
        rayon = Carte.CELL_SIZE // 4
        pygame.draw.circle(Carte.screen, self.color, centre, rayon)


class LesPokemons:
    def __init__(self):
        self.color = (200, 200, 0)  # Couleur jaune pour les Pokémon
        self.pokemons = {}  # Dictionnaire pour stocker les coordonnées des Pokémon
        # Chargement des coordonnées des Pokémon depuis le fichier Excel
        wb = openpyxl.load_workbook("pokemon_coordinates.xlsx")
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            cell1 = sheet.cell(row, 1)
            cell2 = sheet.cell(row, 2)
            pokemon_name = cell1.value
            x, y = json.loads(f"{cell2.value}")  # Convertir la valeur de la cellule en tuple
            self.pokemons[pokemon_name] = (x, y)

    def positionner(self):
        for pokemon_name, (x, y) in self.pokemons.items():
            rayon = Carte.CELL_SIZE // 8  # Taille du rayon pour les Pokémon (2 fois plus petit que le joueur)
            pygame.draw.circle(Carte.screen, self.color, (x, y), rayon)


# Programme principal

# Création du joueur
rows, cols = random.randint(1, Carte.GRID_SIZE) - 1, random.randint(1, Carte.GRID_SIZE) - 1  # La position du joueur
joueur = Joueur(rows, cols)  # Le joueur commence dans une position aléatoire

# Création de pokemon
pokemon = LesPokemons()
pokemon.positionner()
# Boucle principale du jeu
running = True
while running:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
        elif event.type == pygame.KEYDOWN:
            if event.key == pygame.K_UP:
                if event.mod & pygame.K_LEFT:
                    joueur.deplacer("haut-gauche")
                elif event.mod & pygame.K_RIGHT:
                    joueur.deplacer("haut-droite")
                else:
                    joueur.deplacer("haut")
            elif event.key == pygame.K_DOWN:
                if event.mod & pygame.K_LEFT:
                    joueur.deplacer("bas-gauche")
                elif event.mod & pygame.K_RIGHT:
                    joueur.deplacer("bas-droite")
                else:
                    joueur.deplacer("bas")
            elif event.key == pygame.K_LEFT:
                joueur.deplacer("gauche")
            elif event.key == pygame.K_RIGHT:
                joueur.deplacer("droite")

    # Dessiner la carte
    Carte.dessiner_grille()

    # Dessiner le joueur
    joueur.dessiner()

    # Rafraîchir l'écran
    pygame.display.flip()

# Quitter Pygame
pygame.quit()
